import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side
from io import BytesIO

st.title("Обработка main + update Excel")

uploaded_main = st.file_uploader("Загрузите MAIN файл", type=["xlsx"])
uploaded_upd = st.file_uploader("Загрузите UPDATE файл", type=["xlsx"])

if uploaded_main and uploaded_upd and st.button("Запустить обработку"):
    # Загружаем файлы во временные копии в памяти
    main_data = uploaded_main.read()
    upd_data = uploaded_upd.read()

    # Загружаем обновление как DataFrame
    upd = pd.read_excel(BytesIO(upd_data), sheet_name=0, header=None)
    upd.columns = ["ФИО", "M"]
    upd = upd[upd["ФИО"].notna() & upd["M"].notna()]

    # Параметры
    START_ROW = 15
    COL_NUM = 2
    COL_FIO = 3
    COL_L = 12
    COL_M = 13
    COL_N = 14
    COL_O = 15

    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_font = Font(color="FF0000")
    red_border = Border(bottom=Side(style="medium", color="FF0000"))

    def extract_main_fio(fio_full):
        parts = fio_full.strip().upper().split()
        fam = parts[0]
        ini = ""
        if len(parts) > 1: ini = parts[1][0] + "."
        if len(parts) > 2: ini += parts[2][0] + "."
        return fam, ini

    def extract_update_fio(fio_upd):
        parts = fio_upd.strip().upper().split()
        fam = parts[0]
        ini = "".join(parts[1:])
        return fam, ini

    upd[["Fam", "Ini"]] = upd["ФИО"].apply(lambda x: pd.Series(extract_update_fio(x)))

    # Загружаем main с форматированием
    wb = load_workbook(BytesIO(main_data))
    ws = wb.active

    # Построение row_map
    row_map = {}
    current_row = START_ROW

    while True:
        fio_val = ws.cell(row=current_row, column=COL_FIO).value
        if fio_val is None: break
        fam, ini = extract_main_fio(str(fio_val))
        row_map.setdefault(fam, []).append({"row": current_row, "Ini": ini})
        current_row += 1

    last_row = current_row - 1

    update_set = set((r["Fam"], r["Ini"]) for _, r in upd.iterrows())

    # Основной цикл
    for _, r in upd.iterrows():
        fam, ini, inc = r["Fam"], r["Ini"], r["M"]

        if fam in row_map:
            matches = [x for x in row_map[fam] if x["Ini"] == ini]

            if len(matches) == 1:
                row = matches[0]["row"]
                Lp = ws.cell(row=row, column=COL_L).value or 0
                Np = ws.cell(row=row, column=COL_N).value or 0
                ws.cell(row=row, column=COL_L).value = Lp + inc
                ws.cell(row=row, column=COL_M).value = inc
                tax = inc * 0.12
                ws.cell(row=row, column=COL_O).value = tax
                ws.cell(row=row, column=COL_N).value = Np + tax

            else:
                for x in row_map[fam]:
                    ws.cell(row=x["row"], column=COL_FIO).font = red_font
                    ws.cell(row=x["row"], column=COL_FIO).border = red_border

        else:
            last_row += 1
            ws.cell(row=last_row, column=COL_FIO).value = r["ФИО"]
            ws.cell(row=last_row, column=COL_L).value = inc
            ws.cell(row=last_row, column=COL_M).value = inc
            ws.cell(row=last_row, column=COL_N).value = inc * 0.12
            ws.cell(row=last_row, column=COL_O).value = inc * 0.12
            ws.cell(row=last_row, column=COL_FIO).fill = green_fill

    for fam, rows in row_map.items():
        for x in rows:
            if (fam, x["Ini"]) not in update_set:
                ws.cell(row=x["row"], column=COL_M).value = 0

    for i, r in enumerate(range(START_ROW, last_row + 1), start=1):
        ws.cell(row=r, column=COL_NUM).value = i

    # Сохраняем книгу в память
    result = BytesIO()
    wb.save(result)
    result.seek(0)

    st.success("Готово!")
    st.download_button(
        label="💾 Скачать обновлённый Excel",
        data=result,
        file_name="main_updated.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
